"""Small Excel readers shared by backend validation/export code."""

from __future__ import annotations

from pathlib import Path

import pandas as pd


def read_excel_sheets(path: Path, *, header: int | None = 0) -> list[tuple[str, pd.DataFrame]]:
    """Read all sheets from XLS/XLSX.

    For ``.xls`` we avoid ``pandas``' xlrd adapter because modern pandas requires
    xlrd>=2 while dedoc pins xlrd<2 in the shared environment. Direct xlrd usage
    is enough for our table-like validation/export reads and works with both.
    """

    if path.suffix.casefold() == ".xls":
        return _read_xls_sheets(path, header=header)

    if header is None:
        return _read_xlsx_sheets_raw(path)

    excel = pd.ExcelFile(path, engine="openpyxl")
    return [(sheet_name, excel.parse(sheet_name, header=header)) for sheet_name in excel.sheet_names]


def _read_xlsx_sheets_raw(path: Path) -> list[tuple[str, pd.DataFrame]]:
    """Read all cell values without pandas parse overhead (validation hot path)."""

    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets: list[tuple[str, pd.DataFrame]] = []
    try:
        for worksheet in workbook.worksheets:
            rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
            sheets.append((worksheet.title, pd.DataFrame(rows)))
    finally:
        workbook.close()
    return sheets


def _read_xls_sheets(path: Path, *, header: int | None) -> list[tuple[str, pd.DataFrame]]:
    try:
        import xlrd
    except ImportError as exc:
        raise RuntimeError("Для чтения .xls файлов установите xlrd") from exc

    workbook = xlrd.open_workbook(path)
    sheets: list[tuple[str, pd.DataFrame]] = []
    for sheet in workbook.sheets():
        rows = [[sheet.cell_value(row_idx, col_idx) for col_idx in range(sheet.ncols)] for row_idx in range(sheet.nrows)]
        sheets.append((sheet.name, _rows_to_frame(rows, header=header)))
    return sheets


def _rows_to_frame(rows: list[list[object]], *, header: int | None) -> pd.DataFrame:
    if header is None:
        return pd.DataFrame(rows)
    if not rows:
        return pd.DataFrame()
    if header < 0 or header >= len(rows):
        raise ValueError(f"Invalid Excel header row: {header}")
    return pd.DataFrame(rows[header + 1 :], columns=rows[header])
