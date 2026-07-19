from __future__ import annotations

from typing import Any

from openpyxl.styles.fills import DEFAULT_EMPTY_FILL
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

from backend.app.pipeline.runner import PipelineTzResult
from backend.app.schemas import TzPackageRead
from backend.app.services.exporting.common import confidence_label, display, package_recpart, total_value_columns
from backend.app.services.exporting.config import (
    ATTR_CODE_COL_TITLE,
    BLOCK_GAP_ROWS,
    CORRECT_LABELS,
    DATA_ROW_HEIGHT,
    EXECUTION_VARIANT_COL_TITLE,
    HEADER_META_ROW_HEIGHT,
    HEADER_ROW1_HEIGHT,
    HEADER_ROW2_HEIGHT,
    HEADER_ROW3_HEIGHT,
    INCORRECT_LABELS,
    PRED_NO_GT_HIGH_FILL,
    PRED_NO_GT_LOW_FILL,
    PRED_WITH_GT_CORRECT_HIGH_FILL,
    PRED_WITH_GT_CORRECT_LOW_FILL,
    PRED_WITH_GT_INCORRECT_HIGH_FILL,
    PRED_WITH_GT_INCORRECT_LOW_FILL,
    RECPART_COL_TITLE,
    RESULTS_AUX_BLOCK_ALIGN,
    RESULTS_AUX_BLOCK_FILL,
    RESULTS_EXCLUDED_ATTR_FILL,
    RESULTS_HEADER_ALIGN,
    RESULTS_HEADER_FILL,
    RESULTS_HEADER_FONT,
    RESULTS_METADATA_ALIGN,
    RESULTS_METADATA_MUTED_FONT,
    RESULTS_MUTED_META_LABELS,
    RESULTS_MUTED_SECTION_TITLES,
    RESULTS_SECTION_TITLE_ALIGN,
    RESULTS_SECTION_TITLE_FILL,
    RESULTS_SECTION_TITLE_FONT,
    RESULTS_SECTION_TITLE_MUTED_FONT,
    SHEET_RESULTS,
    TITLE_ROW_HEIGHT,
    TZ_ID_COL_TITLE,
    UNIT_SUFFIX,
)
from backend.app.services.exporting.models import AttributeColumn, BlockLayout, ExportEvaluation, ResultsLayout
from backend.app.services.exporting.prediction_maps import prediction_maps


def build_results_sheet(
    worksheet,
    *,
    packages: list[TzPackageRead],
    result_by_package: dict[str, PipelineTzResult],
    attrs: list[AttributeColumn],
    gt_maps: tuple[dict[tuple[str, str], Any], dict[tuple[str, str], str | None]] | None,
    evaluation: ExportEvaluation | None,
) -> ResultsLayout:
    header_top = 1
    write_value_headers(worksheet, header_top, attrs)
    row = header_top + value_header_row_count()

    gt_block: BlockLayout | None = None
    if gt_maps is not None:
        gt_value_map, gt_unit_map = gt_maps
        gt_block = append_value_block(
            worksheet,
            title="GROUND TRUTH",
            title_row=row,
            packages=packages,
            attrs=attrs,
            value_map=gt_value_map,
            unit_map=gt_unit_map,
            fill_mode="gt",
        )
        row = next_block_title_row(gt_block.last_data_row)

    pred_value_map, pred_unit_map, high_confidence_map, source_text_map, _ = prediction_maps(
        packages,
        result_by_package,
    )
    pred_block = append_value_block(
        worksheet,
        title="PREDICTIONS",
        title_row=row,
        packages=packages,
        attrs=attrs,
        value_map=pred_value_map,
        unit_map=pred_unit_map,
        high_confidence_map=high_confidence_map,
        labels_by_recpart=evaluation.labels_by_recpart if evaluation is not None else None,
        fill_mode="pred_with_gt" if evaluation is not None else "pred_no_gt",
    )
    row = next_block_title_row(pred_block.last_data_row)

    source_block = append_aligned_attr_block(
        worksheet,
        title="Фрагменты исходного текста",
        title_row=row,
        packages=packages,
        attrs=attrs,
        value_map=source_text_map,
        gray_for_non_extraction=True,
        muted_values=False,
    )
    row = next_block_title_row(source_block.last_data_row)

    labels_block: BlockLayout | None = None
    if evaluation is not None:
        labels_block = append_aligned_attr_block(
            worksheet,
            title="PREDICTION LABELS",
            title_row=row,
            packages=packages,
            attrs=attrs,
            value_map={
                (package_recpart(package), attr.attr_id): evaluation.labels_by_recpart.get(
                    package_recpart(package),
                    {},
                ).get(attr.attr_id)
                for package in packages
                for attr in attrs
            },
            gray_for_non_extraction=True,
            muted_values=True,
            missing_display=None,
        )
        row = next_block_title_row(labels_block.last_data_row)

    confidence_block = append_aligned_attr_block(
        worksheet,
        title="CONFIDENCE",
        title_row=row,
        packages=packages,
        attrs=attrs,
        value_map={key: confidence_label(value) for key, value in high_confidence_map.items()},
        gray_for_non_extraction=True,
        muted_values=True,
    )

    apply_results_column_widths(worksheet)
    worksheet.freeze_panes = "D2"
    return ResultsLayout(
        first_attr_col=4,
        last_col=total_value_columns(attrs),
        header_row=header_top,
        attr_code_row=header_top + 1,
        essential_row=header_top + 2,
        exclude_row=header_top + 3,
        gt_first_row=gt_block.first_data_row if gt_block is not None else None,
        gt_last_row=gt_block.last_data_row if gt_block is not None else None,
        pred_first_row=pred_block.first_data_row,
        pred_last_row=pred_block.last_data_row,
        source_first_row=source_block.first_data_row,
        source_last_row=source_block.last_data_row,
        labels_first_row=labels_block.first_data_row if labels_block is not None else None,
        labels_last_row=labels_block.last_data_row if labels_block is not None else None,
        confidence_first_row=confidence_block.first_data_row,
        confidence_last_row=confidence_block.last_data_row,
    )


def append_value_block(
    worksheet,
    *,
    title: str,
    title_row: int,
    packages: list[TzPackageRead],
    attrs: list[AttributeColumn],
    value_map: dict[tuple[str, str], Any],
    unit_map: dict[tuple[str, str], str | None],
    high_confidence_map: dict[tuple[str, str], bool | None] | None = None,
    labels_by_recpart: dict[str, dict[str, str]] | None = None,
    fill_mode: str,
) -> BlockLayout:
    total_cols = total_value_columns(attrs)
    write_title_row(worksheet, title_row, title, total_cols)
    first_data_row = title_row + 1

    for row, package in enumerate(packages, start=first_data_row):
        recpart = package_recpart(package)
        write_package_identity(worksheet, row, package)
        col_idx = 4
        for attr in attrs:
            key = (recpart, attr.attr_id)
            fill = value_fill(
                attr,
                fill_mode=fill_mode,
                label=(labels_by_recpart or {}).get(recpart, {}).get(attr.attr_id),
                high_confidence=(high_confidence_map or {}).get(key),
            )
            worksheet.cell(row=row, column=col_idx, value=display(value_map.get(key))).fill = fill
            worksheet.cell(row=row, column=col_idx).alignment = RESULTS_AUX_BLOCK_ALIGN
            if attr.has_unit:
                worksheet.cell(row=row, column=col_idx + 1, value=display(unit_map.get(key))).fill = fill
                worksheet.cell(row=row, column=col_idx + 1).alignment = RESULTS_AUX_BLOCK_ALIGN
                col_idx += 2
            else:
                col_idx += 1
        worksheet.row_dimensions[row].height = DATA_ROW_HEIGHT

    return BlockLayout(
        title_row=title_row,
        first_data_row=first_data_row,
        last_data_row=last_data_row(first_data_row=first_data_row, package_count=len(packages)),
    )


def append_aligned_attr_block(
    worksheet,
    *,
    title: str,
    title_row: int,
    packages: list[TzPackageRead],
    attrs: list[AttributeColumn],
    value_map: dict[tuple[str, str], Any],
    gray_for_non_extraction: bool,
    muted_values: bool = False,
    missing_display: Any = "Н/Д",
) -> BlockLayout:
    total_cols = total_value_columns(attrs)
    write_title_row(worksheet, title_row, title, total_cols)
    first_data_row = title_row + 1

    for row, package in enumerate(packages, start=first_data_row):
        recpart = package_recpart(package)
        write_package_identity(worksheet, row, package, muted=muted_values)
        col_idx = 4
        for attr in attrs:
            fill = RESULTS_EXCLUDED_ATTR_FILL if gray_for_non_extraction and not attr.for_extraction else RESULTS_AUX_BLOCK_FILL
            cell = worksheet.cell(
                row=row,
                column=col_idx,
                value=display(value_map.get((recpart, attr.attr_id)), missing_display=missing_display),
            )
            cell.fill = fill
            if muted_values:
                cell.font = RESULTS_METADATA_MUTED_FONT
            cell.alignment = RESULTS_AUX_BLOCK_ALIGN
            if attr.has_unit:
                unit_cell = worksheet.cell(row=row, column=col_idx + 1, value=None)
                unit_cell.fill = fill
                if muted_values:
                    unit_cell.font = RESULTS_METADATA_MUTED_FONT
                unit_cell.alignment = RESULTS_AUX_BLOCK_ALIGN
                col_idx += 2
            else:
                col_idx += 1
        worksheet.row_dimensions[row].height = DATA_ROW_HEIGHT

    return BlockLayout(
        title_row=title_row,
        first_data_row=first_data_row,
        last_data_row=last_data_row(first_data_row=first_data_row, package_count=len(packages)),
    )


def last_data_row(*, first_data_row: int, package_count: int) -> int:
    if package_count <= 0:
        return first_data_row
    return first_data_row + package_count - 1


def next_block_title_row(last_data_row_value: int) -> int:
    return last_data_row_value + BLOCK_GAP_ROWS + 1


def write_title_row(worksheet, row: int, title: str, total_cols: int) -> None:
    title_cell = worksheet.cell(row=row, column=1, value=title)
    title_cell.font = (
        RESULTS_SECTION_TITLE_MUTED_FONT if title in RESULTS_MUTED_SECTION_TITLES else RESULTS_SECTION_TITLE_FONT
    )
    title_cell.fill = RESULTS_SECTION_TITLE_FILL
    title_cell.alignment = RESULTS_SECTION_TITLE_ALIGN
    for col_idx in range(2, total_cols + 1):
        worksheet.cell(row=row, column=col_idx).fill = RESULTS_SECTION_TITLE_FILL
    worksheet.row_dimensions[row].height = TITLE_ROW_HEIGHT
    if total_cols > 1:
        worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)


def write_value_headers(worksheet, top_row: int, attrs: list[AttributeColumn]) -> None:
    write_identity_headers(worksheet, top_row)
    write_attribute_headers(worksheet, top_row, attrs)
    write_metadata_rows(worksheet, top_row, attrs)
    set_header_heights(worksheet, top_row, rows=value_header_row_count())


def write_identity_headers(worksheet, top_row: int) -> None:
    headers = (TZ_ID_COL_TITLE, EXECUTION_VARIANT_COL_TITLE, RECPART_COL_TITLE)
    for col_idx, title in enumerate(headers, start=1):
        header_cell = worksheet.cell(row=top_row, column=col_idx, value=title)
        header_cell.font = RESULTS_HEADER_FONT
        header_cell.fill = RESULTS_HEADER_FILL
        header_cell.alignment = RESULTS_HEADER_ALIGN
        worksheet.cell(row=top_row + 1, column=col_idx).alignment = RESULTS_METADATA_ALIGN
    worksheet.cell(row=top_row + 1, column=1, value=ATTR_CODE_COL_TITLE).alignment = RESULTS_METADATA_ALIGN


def set_header_heights(worksheet, top_row: int, *, rows: int) -> None:
    worksheet.row_dimensions[top_row].height = HEADER_ROW1_HEIGHT
    if rows >= 2:
        worksheet.row_dimensions[top_row + 1].height = HEADER_ROW2_HEIGHT
    if rows >= 3:
        worksheet.row_dimensions[top_row + 2].height = HEADER_ROW3_HEIGHT
    for offset in range(3, rows):
        worksheet.row_dimensions[top_row + offset].height = HEADER_META_ROW_HEIGHT


def value_header_row_count() -> int:
    return 6


def write_metadata_rows(worksheet, top_row: int, attrs: list[AttributeColumn]) -> None:
    meta_specs = [
        ("определяющий", "essential"),
        ("исключен из расчета метрик", "exclude"),
        ("Подсказка для поиска", "hint_srch"),
        ("Подсказка для экстракции", "hint_ext"),
    ]
    for offset, (label, attr_field) in enumerate(meta_specs, start=2):
        row = top_row + offset
        muted_row = label in RESULTS_MUTED_META_LABELS
        write_row_label(worksheet, row, label, muted=muted_row)
        col_idx = 4
        for attr in attrs:
            value = getattr(attr, attr_field)
            cell = worksheet.cell(row=row, column=col_idx, value=value)
            cell.alignment = RESULTS_METADATA_ALIGN
            if muted_row:
                cell.font = RESULTS_METADATA_MUTED_FONT
            if not attr.for_extraction:
                cell.fill = RESULTS_EXCLUDED_ATTR_FILL
            if attr.has_unit:
                unit_cell = worksheet.cell(row=row, column=col_idx + 1, value=None)
                unit_cell.alignment = RESULTS_METADATA_ALIGN
                if muted_row:
                    unit_cell.font = RESULTS_METADATA_MUTED_FONT
                if not attr.for_extraction:
                    unit_cell.fill = RESULTS_EXCLUDED_ATTR_FILL
                col_idx += 2
            else:
                col_idx += 1


def write_attribute_headers(worksheet, top_row: int, attrs: list[AttributeColumn]) -> None:
    col_idx = 4
    for attr in attrs:
        header_cell = worksheet.cell(row=top_row, column=col_idx, value=attr.attr_name)
        header_cell.font = RESULTS_HEADER_FONT
        header_cell.fill = RESULTS_HEADER_FILL
        header_cell.alignment = RESULTS_HEADER_ALIGN
        worksheet.cell(row=top_row + 1, column=col_idx, value=attr.attr_id).alignment = RESULTS_METADATA_ALIGN
        if attr.has_unit:
            unit_header_cell = worksheet.cell(row=top_row, column=col_idx + 1, value=f"{attr.attr_name} ед.из.")
            unit_header_cell.font = RESULTS_HEADER_FONT
            unit_header_cell.fill = RESULTS_HEADER_FILL
            unit_header_cell.alignment = RESULTS_HEADER_ALIGN
            worksheet.cell(row=top_row + 1, column=col_idx + 1, value=f"{attr.attr_id}{UNIT_SUFFIX}").alignment = (
                RESULTS_METADATA_ALIGN
            )
            col_idx += 2
        else:
            col_idx += 1


def write_row_label(worksheet, row: int, label: str, *, muted: bool = False) -> None:
    label_cell = worksheet.cell(row=row, column=1, value=label)
    label_cell.alignment = RESULTS_METADATA_ALIGN
    if muted:
        label_cell.font = RESULTS_METADATA_MUTED_FONT


def write_package_identity(worksheet, row: int, package: TzPackageRead, *, muted: bool = False) -> None:
    values = (package.tz_id, package.execution_variant, package.recpart)
    for col_idx, value in enumerate(values, start=1):
        cell = worksheet.cell(row=row, column=col_idx, value=display(value))
        cell.alignment = RESULTS_AUX_BLOCK_ALIGN
        if muted:
            cell.font = RESULTS_METADATA_MUTED_FONT


def apply_results_column_widths(worksheet) -> None:
    worksheet.column_dimensions["A"].width = 25.42578125
    worksheet.column_dimensions["B"].width = 14
    worksheet.column_dimensions["C"].width = 16.28515625
    worksheet.column_dimensions["D"].width = 22
    worksheet.column_dimensions["E"].width = 13
    worksheet.column_dimensions["F"].width = 13
    worksheet.column_dimensions["G"].width = 24
    worksheet.column_dimensions["H"].width = 22
    for col_idx in range(9, worksheet.max_column + 1):
        worksheet.column_dimensions[get_column_letter(col_idx)].width = 13


def set_defined_names(workbook, layout: ResultsLayout) -> None:
    if layout.first_attr_col > layout.last_col:
        return
    names: list[tuple[str, str]] = [
        ("Значащий_атрибут", results_range_ref(layout, layout.essential_row, layout.essential_row)),
        ("Исключен_из_метрик", results_range_ref(layout, layout.exclude_row, layout.exclude_row)),
        ("Уверенность", results_range_ref(layout, layout.confidence_first_row, layout.confidence_last_row)),
    ]
    if layout.labels_first_row is not None and layout.labels_last_row is not None:
        names.append(("Лейблы", results_range_ref(layout, layout.labels_first_row, layout.labels_last_row)))

    for name, ref in names:
        if name in workbook.defined_names:
            del workbook.defined_names[name]
        workbook.defined_names.add(DefinedName(name=name, attr_text=ref))


def results_range_ref(layout: ResultsLayout, first_row: int, last_row: int) -> str:
    first_col = get_column_letter(layout.first_attr_col)
    last_col = get_column_letter(layout.last_col)
    return f"{SHEET_RESULTS}!${first_col}${first_row}:${last_col}${last_row}"


def results_labels_range_ref(layout: ResultsLayout) -> str | None:
    if layout.labels_first_row is None or layout.labels_last_row is None:
        return None
    return results_range_ref(layout, layout.labels_first_row, layout.labels_last_row)


def value_fill(
    attr: AttributeColumn,
    *,
    fill_mode: str,
    label: str | None,
    high_confidence: bool | None,
):
    if not attr.for_extraction:
        return RESULTS_EXCLUDED_ATTR_FILL
    if fill_mode == "gt":
        return DEFAULT_EMPTY_FILL
    if fill_mode == "pred_no_gt":
        if high_confidence is True:
            return PRED_NO_GT_HIGH_FILL
        if high_confidence is False:
            return PRED_NO_GT_LOW_FILL
        return DEFAULT_EMPTY_FILL
    if fill_mode == "pred_with_gt":
        if label in CORRECT_LABELS:
            return PRED_WITH_GT_CORRECT_HIGH_FILL if high_confidence is True else PRED_WITH_GT_CORRECT_LOW_FILL
        if label in INCORRECT_LABELS:
            return PRED_WITH_GT_INCORRECT_HIGH_FILL if high_confidence is True else PRED_WITH_GT_INCORRECT_LOW_FILL
    return DEFAULT_EMPTY_FILL
