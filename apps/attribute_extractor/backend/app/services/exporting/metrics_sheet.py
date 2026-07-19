from __future__ import annotations

from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule

from backend.app.services.exporting.config import (
    HEADER_FONT,
    METRICS_ACCOUNTING_FORMAT,
    METRICS_ACCENT_FILL,
    METRICS_ACCENT_FONT,
    METRICS_BORDER_HAIR,
    METRICS_BORDER_MEDIUM,
    METRICS_BORDER_THIN,
    METRICS_CENTER_ALIGN,
    METRICS_DATETIME_FORMAT,
    METRICS_HEADER_ALIGN,
    METRICS_INTEGER_FORMAT,
    METRICS_INVERTED_FONT,
    METRICS_LABEL_HEADER_ROW_HEIGHT,
    METRICS_PERCENT_FORMAT,
    METRICS_SECTION_ALIGN,
    METRICS_SECTION_FILL,
    METRICS_SUMMARY_LAST_ROW_HEIGHT,
    METRICS_TITLE_ALIGN,
    METRICS_TITLE_FONT,
    METRICS_TITLE_ROW_HEIGHT,
    METRICS_TOP_SUMMARY_LABEL_ALIGN,
    METRICS_TOP_SUMMARY_VALUE_ALIGN,
    METRICS_WARNING_FILL,
    METRICS_WARNING_FONT,
    SHEET_AUTO_METRICS,
    SHEET_METRICS,
    SHEET_RESULTS,
    TITLE_FONT,
    WRAP_CENTER,
    metrics_border,
)
from backend.app.services.exporting.models import ResultsLayout
from backend.app.services.exporting.results_sheet import results_labels_range_ref, results_range_ref


def build_metrics_sheet(
    workbook: Workbook,
    *,
    layout: ResultsLayout,
    pipeline_version: str,
    generated_at: datetime,
    sample_count: int,
) -> None:
    worksheet = workbook.create_sheet(title=SHEET_METRICS)
    apply_metrics_column_widths(worksheet)
    configure_workbook_calculation(workbook)
    populate_metrics_headers(worksheet, pipeline_version=pipeline_version, generated_at=generated_at)
    populate_metrics_labels_table(worksheet, layout=layout)
    populate_metrics_scores_table(worksheet)
    populate_metrics_time_block(worksheet, sample_count=sample_count)
    apply_metrics_v1_styles(worksheet, workbook=workbook, layout=layout)


def apply_metrics_column_widths(worksheet) -> None:
    worksheet.column_dimensions["A"].width = 39.28515625
    worksheet.column_dimensions["B"].width = 32.42578125
    worksheet.column_dimensions["C"].width = 26.7109375
    worksheet.column_dimensions["D"].width = 22.85546875
    worksheet.column_dimensions["E"].width = 22.42578125
    worksheet.column_dimensions["F"].width = 16.5703125
    worksheet.column_dimensions["G"].width = 27
    worksheet.column_dimensions["H"].width = 24
    worksheet.column_dimensions["I"].width = 32.140625


def configure_workbook_calculation(workbook: Workbook) -> None:
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"


def populate_metrics_headers(worksheet, *, pipeline_version: str, generated_at: datetime) -> None:
    worksheet.merge_cells("A1:B1")
    worksheet.merge_cells("A11:I11")

    worksheet["A1"].value = "Результаты оценки"
    worksheet["A1"].font = TITLE_FONT
    worksheet["A1"].alignment = WRAP_CENTER

    worksheet["A2"].value = "Net_Effect"
    worksheet["B2"].value = "=B30/B28"
    worksheet["A3"].value = "Доля автоматизации"
    worksheet["B3"].value = "=C22"
    worksheet["A4"].value = "Negative Predictive Value"
    worksheet["B4"].value = "=C23"
    worksheet["A5"].value = "Negative Predictive Value@HC"
    worksheet["B5"].value = "=D23"

    worksheet["A7"].value = "СФОРМИРОВАНО:"
    worksheet["A8"].value = "Дата и время"
    worksheet["B8"].value = generated_at
    worksheet["A9"].value = "Версия пайплайна"
    worksheet["B9"].value = pipeline_version

    worksheet["A11"].value = "РАСЧЕТЫ"
    worksheet["A11"].font = TITLE_FONT

    for cell in ("B3", "B4", "B5"):
        worksheet[cell].number_format = METRICS_PERCENT_FORMAT


def populate_metrics_labels_table(worksheet, *, layout: ResultsLayout) -> None:
    worksheet["A13"].value = "Лейблы"
    worksheet["B13"].value = "Пояснение"
    worksheet["C13"].value = "ВСЕ атрибуты"
    worksheet["D13"].value = "ВСЕ с высокой уверенностью"
    worksheet["E13"].value = "ВСЕ с не высокой уверенностью"
    worksheet["F13"].value = "Определяющие"
    worksheet["G13"].value = "Определяющие с высокой уверенностью"
    worksheet["H13"].value = "Не определяющие с высокой уверенностью"
    worksheet["I13"].value = "Условный норматив на обработку, секунд"

    rows = [
        (14, "TN", "Верно определено отсутствие атрибута", 20),
        (15, "TP", "Верно найдено и распознано и значение, и единица измерения, если имеется", 5),
        (16, "FP1", "Ошибка в определении атрибута - выдумка", 15),
        (17, "FP2", "Ошибка в значении извлеченного атрибута, либо его единицы измерения", 10),
        (18, "FN", "Пропуск (не смогли извлечь данные атрибута из текста)", 20),
    ]
    essential_range = results_range_ref(layout, layout.essential_row, layout.essential_row)
    exclude_range = results_range_ref(layout, layout.exclude_row, layout.exclude_row)
    labels_range = results_labels_range_ref(layout)
    confidence_range = results_range_ref(layout, layout.confidence_first_row, layout.confidence_last_row)
    if labels_range is None:
        raise ValueError("Labels range is required for Metrics sheet")

    for row, label, description, seconds_norm in rows:
        worksheet.cell(row=row, column=1, value=label)
        worksheet.cell(row=row, column=2, value=description)
        worksheet.cell(
            row=row,
            column=3,
            value=sumproduct_formula(
                f'UPPER(TRIM({exclude_range}))="ЛОЖЬ"',
                f"UPPER(TRIM({labels_range}))=$A{row}",
            ),
        )
        worksheet.cell(
            row=row,
            column=4,
            value=sumproduct_formula(
                f'UPPER(TRIM({exclude_range}))="ЛОЖЬ"',
                f"UPPER(TRIM({labels_range}))=$A{row}",
                f'UPPER(TRIM({confidence_range}))="high"',
            ),
        )
        worksheet.cell(
            row=row,
            column=5,
            value=sumproduct_formula(
                f'UPPER(TRIM({exclude_range}))="ЛОЖЬ"',
                f"UPPER(TRIM({labels_range}))=$A{row}",
                f'UPPER(TRIM({confidence_range}))<>"high"',
            ),
        )
        worksheet.cell(
            row=row,
            column=6,
            value=sumproduct_formula(
                f'UPPER(TRIM({essential_range}))="ИСТИНА"',
                f'UPPER(TRIM({exclude_range}))="ЛОЖЬ"',
                f"UPPER(TRIM({labels_range}))=$A{row}",
            ),
        )
        worksheet.cell(
            row=row,
            column=7,
            value=sumproduct_formula(
                f'UPPER(TRIM({essential_range}))="ИСТИНА"',
                f'UPPER(TRIM({exclude_range}))="ЛОЖЬ"',
                f"UPPER(TRIM({labels_range}))=$A{row}",
                f'UPPER(TRIM({confidence_range}))="high"',
            ),
        )
        worksheet.cell(
            row=row,
            column=8,
            value=sumproduct_formula(
                f'UPPER(TRIM({essential_range}))<>"ИСТИНА"',
                f'UPPER(TRIM({exclude_range}))="ЛОЖЬ"',
                f"UPPER(TRIM({labels_range}))=$A{row}",
                f'UPPER(TRIM({confidence_range}))="high"',
            ),
        )
        worksheet.cell(row=row, column=9, value=seconds_norm)

    worksheet["A19"].value = "Total"
    worksheet["B19"].value = "Всего"
    for col in ("C", "D", "E", "F", "G", "H"):
        worksheet[f"{col}19"] = f"=SUM({col}14:{col}18)"


def populate_metrics_scores_table(worksheet) -> None:
    worksheet["A21"].value = "Метрики"
    worksheet["B21"].value = "Формула"
    worksheet["C21"].value = "все атрибуты"
    worksheet["D21"].value = "все с высокой увер."
    worksheet["E21"].value = "все с не высокой увер."
    worksheet["F21"].value = "определяющие"
    worksheet["G21"].value = "определяющие с высокой увер."
    worksheet["H21"].value = "не определяющие с высокой увер."
    worksheet["I21"].value = " "

    worksheet["A22"].value = "Доля автоматизации (Accuracy)"
    worksheet["B22"].value = "(TP + TN) / (TP + TN + FP + FN) × 100%"
    worksheet["A23"].value = "Negative Predictive Value"
    worksheet["B23"].value = "NPV = TN/ (TN + FN)"
    worksheet["A24"].value = "Precision"
    worksheet["B24"].value = "TP / (TP + FP)"
    worksheet["A25"].value = "Recall"
    worksheet["B25"].value = "TP / (TP + FN)"

    for col in ("C", "D", "E", "F", "G", "H"):
        worksheet[f"{col}22"] = f"=({col}15+{col}14)/SUM({col}14:{col}18)"
        worksheet[f"{col}23"] = f"={col}14/({col}14+{col}18)"
        worksheet[f"{col}24"] = f"={col}15/({col}15+{col}16+{col}17)"
        worksheet[f"{col}25"] = f"={col}15/({col}15+{col}18)"
        worksheet[f"{col}22"].number_format = METRICS_PERCENT_FORMAT
        worksheet[f"{col}23"].number_format = METRICS_PERCENT_FORMAT
        worksheet[f"{col}24"].number_format = METRICS_PERCENT_FORMAT
        worksheet[f"{col}25"].number_format = METRICS_PERCENT_FORMAT


def populate_metrics_time_block(worksheet, *, sample_count: int) -> None:
    worksheet["A27"].value = "Время на обработку одного изделия"
    worksheet["B27"].value = "минут"
    worksheet["A28"].value = "текущее"
    worksheet["B28"].value = 40
    worksheet["A29"].value = "прогнозное"
    worksheet["B29"].value = "=(C15*I15+E14*I14+C16*I16+C17*I17+E18*I18)/60/B32"
    worksheet["A30"].value = "сэкономленное"
    worksheet["B30"].value = "=B28-B29"
    worksheet["A32"].value = "Количество образцов"
    worksheet["B32"].value = sample_count


def apply_metrics_v1_styles(worksheet, *, workbook: Workbook, layout: ResultsLayout) -> None:
    worksheet.row_dimensions[1].height = METRICS_TITLE_ROW_HEIGHT
    worksheet.row_dimensions[5].height = METRICS_SUMMARY_LAST_ROW_HEIGHT
    worksheet.row_dimensions[13].height = METRICS_LABEL_HEADER_ROW_HEIGHT

    for cell in worksheet["A1:B1"][0]:
        cell.fill = METRICS_SECTION_FILL
        cell.font = METRICS_TITLE_FONT
        cell.alignment = METRICS_TITLE_ALIGN

    for row in range(2, 6):
        worksheet.cell(row=row, column=1).alignment = METRICS_TOP_SUMMARY_LABEL_ALIGN
        worksheet.cell(row=row, column=2).alignment = METRICS_TOP_SUMMARY_VALUE_ALIGN
    worksheet["B2"].number_format = METRICS_ACCOUNTING_FORMAT

    worksheet["A7"].font = METRICS_TITLE_FONT
    worksheet["A7"].alignment = METRICS_CENTER_ALIGN
    for row in (8, 9):
        worksheet.cell(row=row, column=1).alignment = METRICS_CENTER_ALIGN
        worksheet.cell(row=row, column=2).alignment = METRICS_CENTER_ALIGN
    worksheet["B8"].number_format = METRICS_DATETIME_FORMAT

    apply_metrics_section_row(worksheet, row=11, start_col=1, end_col=9, bold=True, alignment=METRICS_SECTION_ALIGN)
    apply_metrics_section_row(worksheet, row=13, start_col=1, end_col=9, alignment=METRICS_HEADER_ALIGN)
    apply_metrics_section_row(worksheet, row=21, start_col=1, end_col=8)
    apply_metrics_section_row(worksheet, row=27, start_col=1, end_col=2)

    for coordinate in ("C16", "C17"):
        worksheet[coordinate].fill = METRICS_ACCENT_FILL
    for coordinate in ("E14", "C15", "E18", "C22", "C23", "D23"):
        cell = worksheet[coordinate]
        cell.fill = METRICS_ACCENT_FILL
        cell.font = METRICS_ACCENT_FONT
    apply_metrics_warning_rules(worksheet)
    apply_metrics_g18_warning_style(worksheet, workbook=workbook, layout=layout)
    apply_metrics_borders(worksheet)

    for coordinate in ("B28", "B29", "B30", "B32"):
        worksheet[coordinate].number_format = METRICS_INTEGER_FORMAT


def apply_metrics_g18_warning_style(worksheet, *, workbook: Workbook, layout: ResultsLayout) -> None:
    if count_metrics_g18_warning(workbook[SHEET_RESULTS], layout) <= 0:
        return
    cell = worksheet["G18"]
    cell.fill = METRICS_WARNING_FILL
    cell.font = METRICS_WARNING_FONT


def count_metrics_g18_warning(results_worksheet, layout: ResultsLayout) -> int:
    if layout.labels_first_row is None or layout.labels_last_row is None:
        return 0

    label_rows = range(layout.labels_first_row, layout.labels_last_row + 1)
    confidence_rows = range(layout.confidence_first_row, layout.confidence_last_row + 1)
    count = 0
    for col in range(layout.first_attr_col, layout.last_col + 1):
        if _is_excluded_from_metrics(results_worksheet.cell(layout.exclude_row, col).value):
            continue
        if _is_essential_attribute(results_worksheet.cell(layout.essential_row, col).value):
            continue
        labels = [
            str(results_worksheet.cell(row, col).value or "").strip().upper() for row in label_rows
        ]
        if "FN" not in labels:
            continue
        confidences = [
            str(results_worksheet.cell(row, col).value or "").strip().lower() for row in confidence_rows
        ]
        if "high" not in confidences:
            continue
        count += 1
    return count


def _is_excluded_from_metrics(value) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    return str(value).strip().upper() in {"TRUE", "ИСТИНА", "1"}


def _is_essential_attribute(value) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    return str(value).strip().upper() in {"TRUE", "ИСТИНА", "1"}


def apply_metrics_borders(worksheet) -> None:
    medium = METRICS_BORDER_MEDIUM
    thin = METRICS_BORDER_THIN
    hair = METRICS_BORDER_HAIR

    worksheet["A1"].border = metrics_border(left=medium, right=medium, top=medium, bottom=medium)
    worksheet["B1"].border = metrics_border(right=medium, top=medium, bottom=medium)

    for row in (2, 3):
        worksheet.cell(row=row, column=1).border = metrics_border(
            left=medium, right=thin, top=thin, bottom=thin
        )
        worksheet.cell(row=row, column=2).border = metrics_border(
            left=thin, right=medium, top=thin, bottom=thin
        )
    worksheet["A4"].border = metrics_border(left=medium, right=thin, top=thin)
    worksheet["B4"].border = metrics_border(left=thin, right=medium, top=thin)
    worksheet["A5"].border = metrics_border(left=medium, right=thin, top=thin, bottom=medium)
    worksheet["B5"].border = metrics_border(left=thin, right=medium, top=thin, bottom=medium)

    for row in (8, 9):
        for col in (1, 2):
            worksheet.cell(row=row, column=col).border = metrics_border(
                left=thin, right=thin, top=thin, bottom=thin
            )

    apply_metrics_table_borders(
        worksheet,
        header_row=13,
        first_data_row=14,
        last_data_row=18,
        total_row=19,
    )
    apply_metrics_table_borders(
        worksheet,
        header_row=21,
        first_data_row=22,
        last_data_row=24,
        total_row=25,
        last_col=8,
    )
    apply_metrics_time_block_borders(worksheet)


def apply_metrics_table_borders(
    worksheet,
    *,
    header_row: int,
    first_data_row: int,
    last_data_row: int,
    total_row: int | None = None,
    first_col: int = 1,
    last_col: int = 9,
) -> None:
    thin = METRICS_BORDER_THIN
    hair = METRICS_BORDER_HAIR

    for col in range(first_col, last_col + 1):
        cell = worksheet.cell(row=header_row, column=col)
        if col == first_col:
            cell.border = metrics_border(left=thin, right=hair, top=thin, bottom=hair)
        elif col == last_col:
            cell.border = metrics_border(left=hair, right=thin, top=thin, bottom=hair)
        else:
            cell.border = metrics_border(left=hair, right=hair, top=thin, bottom=hair)

    for row in range(first_data_row, last_data_row + 1):
        for col in range(first_col, last_col + 1):
            cell = worksheet.cell(row=row, column=col)
            if col == first_col:
                cell.border = metrics_border(left=thin, right=hair, top=hair, bottom=hair)
            elif col == last_col:
                cell.border = metrics_border(left=hair, right=thin, top=hair, bottom=hair)
            else:
                cell.border = metrics_border(left=hair, right=hair, top=hair, bottom=hair)

    if total_row is None:
        return

    for col in range(first_col, last_col + 1):
        cell = worksheet.cell(row=total_row, column=col)
        if col == first_col:
            cell.border = metrics_border(left=thin, right=hair, top=hair, bottom=thin)
        elif col == last_col:
            cell.border = metrics_border(left=hair, right=thin, top=hair, bottom=thin)
        else:
            cell.border = metrics_border(left=hair, right=hair, top=hair, bottom=thin)


def apply_metrics_time_block_borders(worksheet) -> None:
    thin = METRICS_BORDER_THIN
    hair = METRICS_BORDER_HAIR

    worksheet["A27"].border = metrics_border(left=thin, right=hair, top=thin, bottom=hair)
    worksheet["B27"].border = metrics_border(left=hair, right=thin, top=thin, bottom=hair)

    for row in (28, 29):
        worksheet.cell(row=row, column=1).border = metrics_border(
            left=thin, right=hair, top=hair, bottom=hair
        )
        worksheet.cell(row=row, column=2).border = metrics_border(
            left=hair, right=thin, top=hair, bottom=hair
        )

    worksheet["A30"].border = metrics_border(left=thin, right=hair, top=hair, bottom=thin)
    worksheet["B30"].border = metrics_border(left=hair, right=thin, top=hair, bottom=thin)
    worksheet["B32"].border = metrics_border(left=hair, right=thin, top=thin, bottom=thin)


def apply_metrics_warning_rules(worksheet) -> None:
    worksheet.conditional_formatting.add(
        "G18",
        CellIsRule(operator="greaterThan", formula=["0"], fill=METRICS_WARNING_FILL, font=METRICS_WARNING_FONT),
    )


def apply_metrics_section_row(
    worksheet,
    *,
    row: int,
    start_col: int,
    end_col: int,
    bold: bool = False,
    alignment=None,
) -> None:
    for col in range(start_col, end_col + 1):
        cell = worksheet.cell(row=row, column=col)
        cell.fill = METRICS_SECTION_FILL
        cell.font = METRICS_TITLE_FONT if bold else METRICS_INVERTED_FONT
        if alignment is not None:
            cell.alignment = alignment


def sumproduct_formula(*conditions: str) -> str:
    terms = "*".join(f"--({condition})" for condition in conditions)
    return f"=SUMPRODUCT({terms})"


def build_auto_metrics_sheet(workbook: Workbook, metrics: dict[str, Any]) -> None:
    worksheet = workbook.create_sheet(title=SHEET_AUTO_METRICS)
    worksheet.column_dimensions["A"].width = 28
    worksheet.column_dimensions["B"].width = 60
    worksheet["A1"].value = "metric"
    worksheet["B1"].value = "value"
    worksheet["A1"].font = HEADER_FONT
    worksheet["B1"].font = HEADER_FONT
    surface = metrics.get("extraction_surface") or {}
    rows = [
        ("accuracy", surface.get("accuracy")),
        ("n_cases", surface.get("n_cases")),
        ("error_count", surface.get("error_count")),
        ("high_confidence.n", (surface.get("high_confidence") or {}).get("n")),
        ("high_confidence.accuracy", (surface.get("high_confidence") or {}).get("accuracy")),
        ("low_confidence.n", (surface.get("low_confidence") or {}).get("n")),
        ("low_confidence.accuracy", (surface.get("low_confidence") or {}).get("accuracy")),
    ]
    for row_idx, (key, value) in enumerate(rows, start=2):
        worksheet.cell(row=row_idx, column=1, value=key)
        worksheet.cell(row=row_idx, column=2, value=value)
