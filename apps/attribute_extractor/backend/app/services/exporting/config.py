from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.colors import Color


UNIT_SUFFIX = " ед.из."

SHEET_RESULTS = "Results"
SHEET_METRICS = "Metrics"
SHEET_AUTO_METRICS = "Auto_Metrics"
EXPORT_FORMAT_VERSION = 7

BLOCK_GAP_ROWS = 1
TITLE_ROW_HEIGHT = 21.95
HEADER_ROW1_HEIGHT = 45
HEADER_ROW2_HEIGHT = 20.1
HEADER_ROW3_HEIGHT = 20.1
HEADER_META_ROW_HEIGHT = 18
DATA_ROW_HEIGHT = 18
METRICS_PERCENT_FORMAT = "0.00%"

TZ_ID_COL_TITLE = "ИД документа"
ATTR_CODE_COL_TITLE = "код атрибута"
EXECUTION_VARIANT_COL_TITLE = "Вариант исполнения"
RECPART_COL_TITLE = "RECPart"

BLUE_FILL = PatternFill(patternType="solid", fgColor="FFBDD7EE")
DARK_BLUE_FILL = PatternFill(patternType="solid", fgColor="FF5B9BD5")
GREEN_FILL = PatternFill(patternType="solid", fgColor="FFC6EFCE")
DARK_GREEN_FILL = PatternFill(patternType="solid", fgColor="FF70AD47")
RED_FILL = PatternFill(patternType="solid", fgColor="FFFFC7CE")
DARK_RED_FILL = PatternFill(patternType="solid", fgColor="FFC65911")
GRAY_FILL = PatternFill(patternType="solid", fgColor="FFC0C0C0")
HEADER_FONT = Font(bold=True)
TITLE_FONT = Font(bold=True)
WRAP_CENTER = Alignment(wrap_text=True, horizontal="center", vertical="center")
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
DATA_TOP = Alignment(wrap_text=False, vertical="top")
TITLE_ALIGN = Alignment(horizontal="left", vertical="center")

# Results sheet style groups (v1).
RESULTS_MUTED_TEXT_COLOR = Color(theme=0, tint=-0.5)
RESULTS_HEADER_FILL = PatternFill(patternType="solid", fgColor=Color(theme=0, tint=-0.1499984740745262))
RESULTS_HEADER_FONT = Font(bold=True)
RESULTS_HEADER_ALIGN = Alignment(wrap_text=True, horizontal="center", vertical="center")

RESULTS_METADATA_ALIGN = Alignment(wrap_text=True, vertical="top")
RESULTS_METADATA_MUTED_FONT = Font(color=RESULTS_MUTED_TEXT_COLOR)

RESULTS_SECTION_TITLE_FILL = PatternFill(patternType="solid", fgColor=Color(theme=0, tint=-0.0499893185216834))
RESULTS_SECTION_TITLE_FONT = Font(bold=True)
RESULTS_SECTION_TITLE_MUTED_FONT = Font(bold=True, color=RESULTS_MUTED_TEXT_COLOR)
RESULTS_SECTION_TITLE_ALIGN = Alignment(horizontal="left", vertical="center")
RESULTS_MUTED_SECTION_TITLES = frozenset({"PREDICTION LABELS", "CONFIDENCE"})
RESULTS_MUTED_META_LABELS = frozenset({"определяющий", "исключен из расчета метрик"})

RESULTS_AUX_BLOCK_FILL = PatternFill(patternType=None)
RESULTS_AUX_BLOCK_ALIGN = Alignment(vertical="top")

RESULTS_EXCLUDED_ATTR_FILL = GRAY_FILL

# Metrics sheet style groups (v1).
METRICS_SECTION_FILL = PatternFill(patternType="solid", fgColor=Color(theme=0, tint=-0.0499893185216834))
METRICS_ACCENT_FILL = PatternFill(patternType="solid", fgColor=Color(theme=6, tint=0.7999816888943144))
METRICS_WARNING_FILL = PatternFill(patternType="solid", fgColor=Color(theme=5, tint=0.7999816888943144))
METRICS_TITLE_FONT = Font(bold=True)
METRICS_INVERTED_FONT = Font(color=Color(theme=1))
METRICS_ACCENT_FONT = Font(bold=True, color=Color(theme=1))
METRICS_WARNING_FONT = Font(bold=True, color=Color(theme=1))
METRICS_TITLE_ALIGN = Alignment(wrap_text=True, horizontal="center", vertical="center")
METRICS_SECTION_ALIGN = Alignment(wrap_text=True, horizontal="left", vertical="center")
METRICS_HEADER_ALIGN = Alignment(wrap_text=True, vertical="top")
METRICS_TOP_SUMMARY_LABEL_ALIGN = Alignment(wrap_text=True, horizontal="center", vertical="center")
METRICS_TOP_SUMMARY_VALUE_ALIGN = Alignment(wrap_text=True, horizontal="right", vertical="center")
METRICS_CENTER_ALIGN = Alignment(wrap_text=True, horizontal="center", vertical="center")
METRICS_INTEGER_FORMAT = "0"
METRICS_ACCOUNTING_FORMAT = '_-* #,##0.00_-;\\-* #,##0.00_-;_-* "-"??_-;_-@_-'
METRICS_DATETIME_FORMAT = "m/d/yy h:mm"
METRICS_TITLE_ROW_HEIGHT = 15.75
METRICS_SUMMARY_LAST_ROW_HEIGHT = 15.75
METRICS_LABEL_HEADER_ROW_HEIGHT = 33.75

METRICS_BORDER_MEDIUM = Side(style="medium")
METRICS_BORDER_THIN = Side(style="thin")
METRICS_BORDER_HAIR = Side(style="hair")


def metrics_border(
    *,
    left: Side | None = None,
    right: Side | None = None,
    top: Side | None = None,
    bottom: Side | None = None,
) -> Border:
    return Border(left=left, right=right, top=top, bottom=bottom)


PRED_WITH_GT_CORRECT_LOW_FILL = GREEN_FILL
PRED_WITH_GT_CORRECT_HIGH_FILL = DARK_GREEN_FILL
PRED_WITH_GT_INCORRECT_LOW_FILL = RED_FILL
PRED_WITH_GT_INCORRECT_HIGH_FILL = DARK_RED_FILL
PRED_NO_GT_LOW_FILL = BLUE_FILL
PRED_NO_GT_HIGH_FILL = DARK_BLUE_FILL

CORRECT_LABELS = frozenset({"TP", "TN"})
INCORRECT_LABELS = frozenset({"FP1", "FP2", "FN"})
