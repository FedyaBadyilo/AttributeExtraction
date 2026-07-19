from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

MANIFEST_COLUMNS = {
    "Идентификатор ЕОС НСИ": "gid",
    "Имя файла с расширением": "pdf_filename",
    "Номер для обработки": "file_priority",
    "Идентификатор варианта исполнения": "variant_execution_id",
}


def read_manifest(path: Path) -> list[dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    header = [str(cell).strip() if cell is not None else "" for cell in next(rows)]
    column_map = {MANIFEST_COLUMNS[name]: header.index(name) for name in MANIFEST_COLUMNS}

    examples: list[dict] = []
    for row in rows:
        if row is None or all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        gid = int(row[column_map["gid"]])
        variant = row[column_map["variant_execution_id"]]
        examples.append(
            {
                "gid": gid,
                "eos_id": gid,
                "pdf_filename": str(row[column_map["pdf_filename"]]).strip(),
                "file_priority": int(row[column_map["file_priority"]]),
                "variant_execution_id": None if variant is None else str(variant).strip(),
            }
        )

    workbook.close()
    return examples
